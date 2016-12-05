import requests
import argparse
import json

from datetime import datetime
from lxml import etree
from bs4 import BeautifulSoup
from openpyxl import Workbook


def get_courses_list(courses_count):
    xml_feed_url = 'https://www.coursera.org/sitemap~www~courses.xml'
    response = requests.get(xml_feed_url)
    root = etree.fromstring(response.text.encode('utf-8'))
    namespace = list(root.nsmap.values())[0]
    loc_tags = root.xpath('//ns:loc', namespaces={'ns': namespace})
    courses_list = []
    for url in loc_tags:
        courses_list.append(url.text)
    return courses_list[:courses_count]


def get_course_info(course_url):
    def get_element_text(elem):
        return elem.text if elem else 'No information'

    response = requests.get(course_url)
    soup = BeautifulSoup(response.text, 'lxml')

    course_info = {}
    json_elem = soup.find('script', {'type': 'application/ld+json'})
    if json_elem:
        json_course_data = json.loads(json_elem.text)
        has_course_instance = json_course_data['hasCourseInstance'][0]
        if 'startDate' in has_course_instance and 'endDate' in has_course_instance:
            course_info['start date'] = has_course_instance['startDate']
            end_date = has_course_instance['endDate']
            start_date = datetime.strptime(course_info['start date'], '%Y-%m-%d')
            end_date = datetime.strptime(end_date, '%Y-%m-%d')
            course_info['weeks count'] = (end_date - start_date).days / 7 - 1
        else:
            course_info['start date'] = course_info['weeks count'] = 'No information'
    else:
        course_info['start date'] = course_info['weeks count'] = 'No information'

    course_info['url'] = course_url
    course_info['title'] = get_element_text(soup.find('h1', {'class': 'course-name-text'}))
    course_info['language'] = get_element_text(soup.find('div', {'class': 'language-info'}))
    course_info['rating'] = get_element_text(soup.find('div', {'class': 'ratings-text'}))

    return course_info


def output_courses_info_to_xlsx(filepath, courses_info):
    wb = Workbook()
    ws = wb.create_sheet("Courses info", 0)

    str_list = ['title', 'start date', 'weeks count', 'language', 'rating', 'url']

    for row in range(len(courses_info) + 1):
        for col in range(len(str_list)):
            if row == 0:
                ws.cell(row=row+1, column=col+1, value=str_list[col].upper())
            else:
                ws.cell(row=row+1, column=col+1, value=courses_info[row - 1][str_list[col]])

    wb.save(filepath)
    print('Courses information is saved to the file \"{0}\"!'.format(filepath))


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='The script obtains Coursera courses information and unloads it to the xlsx-file'
    )
    parser.add_argument('xlsx_file_name', help='xlsx-file name')
    parser.add_argument('count', help='Courses count', type=int)
    args = parser.parse_args()

    courses_count = args.count
    courses_url_list = get_courses_list(courses_count)

    courses_info = []
    for num, course_url in enumerate(courses_url_list):
        print('[{0}/{1}] Loading from {2}'.format(num + 1, courses_count, course_url))
        courses_info.append(get_course_info(course_url))

    output_courses_info_to_xlsx(args.xlsx_file_name, courses_info)
